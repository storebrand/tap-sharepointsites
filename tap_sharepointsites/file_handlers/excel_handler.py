"""Handle Excel files."""

import logging
import re
import tempfile
from tap_sharepointsites.utils import snakecase
import openpyxl

LOGGER = logging.getLogger(__name__)


class ExcelHandler:
    """Handle Excel files."""

    def __init__(self, textcontent, clean_colnames=False):
        """Initialize ExcelHandler."""
        self.xlsheet = self._load_workbook(textcontent)
        self.clean_colnames = clean_colnames

    def _load_workbook(self, textcontent):
        """Load workbook from textcontent."""
        with tempfile.NamedTemporaryFile(mode="wb", suffix=".xlsx") as temp:
            temp.write(textcontent)
            temp.flush()
            workbook = openpyxl.load_workbook(temp.name, read_only=True)
            worksheets = workbook.worksheets
            active_sheet = worksheets[0]
            return active_sheet
            # self.xlsheet = active_sheet

    def get_row_iterator(self):
        """Return a generator of rows."""
        yield from self.generator_wrapper(self.xlsheet)

    @property
    def fieldnames(self):
        """Return fieldnames."""
        return [c.value for c in self.xlsheet[1]]

    def generator_wrapper(self, reader):
        """Wrap a reader in a generator."""
        header_row = None
        for row in reader:
            to_return = {}
            if header_row is None:
                header_row = row
                continue

            for index, cell in enumerate(row):
                header_cell = header_row[index]

                formatted_key = header_cell.value


                # if self.clean_colnames:
                #     self.logger.info("cleaning colnames")
                #     formatted_key = snakecase(formatted_key)

                if not formatted_key:
                    formatted_key = ""  # default to empty string for key


                to_return[formatted_key] = (
                    str(cell.value) if cell.value is not None else ""
                )

            yield to_return
